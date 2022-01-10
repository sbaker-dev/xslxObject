from .SheetData import SheetData

from openpyxl.utils import get_column_letter
from miscSupports import validate_path
from openpyxl import load_workbook
from typing import List, Union
from pathlib import Path


class XlsxObject:
    """
    This class is designed to create a Xlsx object using openpyxl to parse in the information from an xlsx file.
    """

    def __init__(self, read_file: Union[str, Path], file_headers=True):
        """
        This object takes a read file directory as its core argument. By default file headers are turned on, but if a
        file doesn't have any file headers users can turn file headers of.

        This object has the following attributes:

        file_name: The file name of the read file minus any file extension

        sheet_column_lengths: The number of columns of data that exist with each given sheet. Keep in mind, this
        operation assumes that even if you don't have column headers, that the first row of every column does have
        content in it as this determines the length.

        sheet_row_lengths: The number of rows of data for the columns within a given sheet. Keep in mind, this operation
        assumes equal length columns and only takes the length of the first column in each sheet as the true value
        for all other columns within the sheet.

        sheet_headers: If headers are set to true, this will isolate the first row of each given column for each given
        sheet and use these values as the header value for a given column in a given sheet.

        sheet_data: This contains the data from all the sheets in a sheet-column-row format.

        :param read_file: The xlsx file path you want to read in to an object
        :type read_file: str | Path

        :param file_headers: If the xlsx file has file headers or not
        :type file_headers: bool
        """

        self._read_file = validate_path(read_file)
        self._workbook = load_workbook(self._read_file)
        self._file_headers = file_headers

        self.file_name = self._read_file.stem
        self.sheet_names = self._set_sheet_names()
        self.sheet_col_count = [sheet.max_column for sheet in self._workbook.worksheets]
        self.sheet_row_count = [sheet.max_row for sheet in self._workbook.worksheets]
        self.sheet_headers = self._set_sheet_header_list()
        self.sheet_data = self._set_sheet_data()

    def __repr__(self):
        """Human readable print"""
        return f"{self.file_name}.xlsx with {len(self.sheet_names)} sheets"

    def __getitem__(self, item):
        """Extract the data """
        if isinstance(item, int):
            return self.sheet_data[item]
        else:
            raise TypeError(f"Getting sheet data via __getitem__ requires an item yet was passed {type(item)}")

    def _set_sheet_names(self) -> List[str]:
        """
        This extracts the sheets titles from the xlsx workbook
        """
        return [sheet.title for sheet in self._workbook.worksheets]

    def _set_sheet_header_list(self) -> List[List[str]]:
        """
        Isolates headers if they exist, else creates dummy header names for each sheet in workbook
        """
        if self._file_headers:
            sheet_headers = [[sheet[f"{get_column_letter(i)}{1}"].value for i in range(1, sheet_length + 1)]
                             for sheet_length, sheet in zip(self.sheet_col_count, self._workbook.worksheets)]

        else:
            sheet_headers = [[f"Var{i}" for i in range(1, sheet_length)] for sheet_length in self.sheet_col_count]
        return sheet_headers

    def _set_sheet_data(self) -> List[SheetData]:
        """
        Iterator that will work through the sheets by using the column and row lengths, isolating all the content
        within a given sheet. This means that the end result is a nested list of sheet-column-row.
        """
        return [self._set_data(sheet, sheet_index) for sheet_index, sheet in enumerate(self._workbook.worksheets)]

    def _set_data(self, sheet, sheet_index: int) -> SheetData:
        """
        This sets the data for a given sheet by taking the row and column lengths and then iterating through the sheets
        columns and rows by using range indexing.

        NOTE
        ----
        openpyxl requires base 1 not base 0 hence range
        """
        # Set row count based on if headers are include so that headers are not within data rows
        if self._file_headers:
            row_start = 2
        else:
            row_start = 1
        row_end = self.sheet_row_count[sheet_index] + 1

        # Extract the data from the sheet
        sheet_data = [[sheet[f"{get_column_letter(col_i)}{row_i}"].value for row_i in range(row_start, row_end)]
                      for col_i in range(1, self.sheet_col_count[sheet_index] + 1)]

        # Return the sheet data type SheetData
        return SheetData(self.sheet_names[sheet_index], self.sheet_headers[sheet_index], sheet_data)
